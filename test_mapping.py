import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
MAPPING_FILE = ROOT / "mapping_master.json"
DATA_FILE = ROOT / "data.bk"


def is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    if isinstance(value, list) and len(value) == 0:
        return True
    return False


def normalize_text(text: str) -> str:
    return " ".join(str(text).strip().lower().split())


def to_attr_dict(client: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for item in client.get("ClientAttributes", []):
        key = item.get("Key")
        if key is not None:
            out[str(key)] = item.get("Value")
    return out


def get_assets_by_code(client: Dict[str, Any], code: str) -> List[Dict[str, Any]]:
    return [a for a in client.get("ClientAssets", []) if a.get("Code") == code]


def asset_props_to_dict(asset: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for item in asset.get("AssetProperties", []):
        key = item.get("Key")
        if key is not None:
            out[str(key)] = item.get("Value")
    return out


def resolve_data_path(client: Dict[str, Any], path: str) -> Any:
    m = re.fullmatch(r"Clients\[0\]\.ClientAttributes\['([^']+)'\]", path)
    if m:
        attrs = to_attr_dict(client)
        return attrs.get(m.group(1))

    m = re.fullmatch(
        r"Clients\[0\]\.ClientAssets\[Code='([^']+)'\]\.AssetProperties\['([^']+)'\]",
        path,
    )
    if m:
        code, key = m.groups()
        assets = get_assets_by_code(client, code)
        if not assets:
            return None
        props = asset_props_to_dict(assets[0])
        return props.get(key)

    m = re.fullmatch(r"Clients\[0\]\.ClientAssets\[Code='([^']+)'\]\[\]", path)
    if m:
        return get_assets_by_code(client, m.group(1))

    m = re.fullmatch(
        r"Clients\[0\]\.ClientAssets\[Code='([^']+)'\]\[\*\]\.AssetProperties\['([^']+)'\]",
        path,
    )
    if m:
        code, key = m.groups()
        values = []
        for asset in get_assets_by_code(client, code):
            props = asset_props_to_dict(asset)
            if key in props:
                values.append(props[key])
        return values

    return None


def find_year_columns(ws) -> Dict[int, int]:
    year_cols: Dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=40):
        local: Dict[int, int] = {}
        for idx, cell in enumerate(row, start=1):
            v = cell.value
            if v is None:
                continue
            s = str(v)
            hits = re.findall(r"(20\d{2})", s)
            for h in hits:
                year = int(h)
                # Keep first matched column for each year in the header row,
                # avoid overriding with "So sanh 2024/2023" columns.
                if year not in local:
                    local[year] = idx
        if len(local) >= 2:
            return local
    return year_cols


def row_matches_label(row_values: List[Any], label: str) -> bool:
    row_text = " | ".join(str(v) for v in row_values if isinstance(v, str) and v.strip())
    row_norm = normalize_text(row_text)

    label_raw = label.strip()
    label_base = re.sub(r"\s*\([^)]*\)\s*", " ", label_raw).strip()
    label_base_norm = normalize_text(label_base)

    if label_base_norm and label_base_norm in row_norm:
        return True

    code_match = re.search(r"\((\d+)\)", label_raw)
    if code_match:
        code = code_match.group(1)
        for v in row_values:
            if v is None:
                continue
            if str(v).strip() == code:
                return True

    label_norm = normalize_text(label_raw)
    return bool(label_norm and label_norm in row_norm)


def extract_numeric_from_row(row_values: List[Any]) -> List[float]:
    nums = []
    for v in row_values:
        if isinstance(v, (int, float)):
            nums.append(float(v))
    return nums


def resolve_xlsm_path(wb, path: str) -> Any:
    m = re.fullmatch(r"([^\[]+)\['([^']+)'\](?:\[year=latest\])?", path)
    if not m:
        return None

    sheet_name, label = m.groups()
    sheet_name = sheet_name.strip()
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    year_cols = find_year_columns(ws)
    latest_year_col: Optional[int] = None
    if year_cols:
        latest_year = max(year_cols.keys())
        latest_year_col = year_cols[latest_year]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        row_values = list(row)
        if not row_matches_label(row_values, label):
            continue

        if latest_year_col is not None and latest_year_col <= len(row_values):
            v = row_values[latest_year_col - 1]
            if isinstance(v, (int, float)):
                return v

        nums = extract_numeric_from_row(row_values)
        if nums:
            return nums[-1]
        return None

    return None


def main() -> None:
    mapping = json.loads(MAPPING_FILE.read_text(encoding="utf-8"))
    data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    client = data["Clients"][0]

    xlsm_file = ROOT / mapping["sources"]["xlsm_hmtd"]["file"]
    wb = load_workbook(xlsm_file, data_only=True, read_only=True)

    total = 0
    ok = 0
    missing = 0

    print("=== Mapping Test Result ===")
    for item in mapping["mappings"]:
        total += 1
        field = item["template_field"]
        status = item.get("status", "")

        if status == "MISSING":
            missing += 1
            print(f"[SKIP-MISSING] {field}")
            continue

        resolved_value = None
        resolved_from = None
        for src in item.get("sources", []):
            src_name = src.get("source")
            src_path = src.get("path", "")
            value = None
            if src_name == "data_bk":
                value = resolve_data_path(client, src_path)
            elif src_name == "xlsm_hmtd":
                value = resolve_xlsm_path(wb, src_path)

            if not is_empty(value):
                resolved_value = value
                resolved_from = f"{src_name}:{src_path}"
                break

        if is_empty(resolved_value):
            print(f"[FAIL] {field} -> no value")
        else:
            ok += 1
            preview = str(resolved_value)
            if len(preview) > 120:
                preview = preview[:117] + "..."
            print(f"[OK] {field} -> {preview} ({resolved_from})")

    print("\n=== Summary ===")
    print(f"Total fields: {total}")
    print(f"Resolved OK: {ok}")
    print(f"Declared missing (skipped): {missing}")
    print(f"Failed unresolved: {total - ok - missing}")


if __name__ == "__main__":
    main()
